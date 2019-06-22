from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import time
import pymysql
import os
from openpyxl import load_workbook

def crawling(user_nick):

    driver.get('https://www.op.gg/') 

    search_name = driver.find_element_by_name('userName')

    search_name.send_keys(user_nick) 
    search_name.submit()

    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')

    UserTierRankInfo = soup.find('div', class_='TierRankInfo')

    DivUserTierRankInfo = UserTierRankInfo.find_all('div')
    TireRank= DivUserTierRankInfo[1].get_text()
    win = int(DivUserTierRankInfo[2].find('span', class_='wins').get_text()[:-1])
    loss = int(DivUserTierRankInfo[2].find('span', class_='losses').get_text()[:-1])

    try:
        sql = """
            INSERT INTO testEntireInfo
                (nickname, rank, wins, losses)
            VALUES
                (%s, %s, %s, %s)
            """
        cursor.execute(sql, (user_nick, TireRank, win, loss))
    except pymysql.IntegrityError:
        sql = """
            UPDATE testEntireInfo SET rank = %s, wins = %s, losses = %s WHERE nickname = %s
            """
        cursor.execute(sql, (TireRank, win, loss, user_nick))


    UserPositionStatContent = soup.find_all('div', class_='PositionStatContent')

    position1 = UserPositionStatContent[0].div.get_text()
    UserPositionStatContentDetail0 = UserPositionStatContent[0].find_all('b')
    RoleRate1 = float(UserPositionStatContentDetail0[0].get_text())
    WinRatio1 = float(UserPositionStatContentDetail0[1].get_text())
    

    position2 = UserPositionStatContent[1].div.get_text()
    if position2 != "?" :
        UserPositionStatContentDetail1 = UserPositionStatContent[1].find_all('b')
        RoleRate2 = float(UserPositionStatContentDetail1[0].get_text())
        WinRatio2 = float(UserPositionStatContentDetail1[1].get_text())

    try:
        sql = """
            INSERT INTO testPosition
                (nickname, position, role_rate, win_rate)
            values
                (%s, %s, %s, %s)
            """
        cursor.execute(sql, (user_nick, position1, RoleRate1, WinRatio1))
        if position2 != '?':
            cursor.execute(sql, (user_nick, position2, RoleRate2, WinRatio2))
    except pymysql.IntegrityError:
        sql = """
            DELETE FROM testPosition
            WHERE nickname = %s
            """
        cursor.execute(sql, user_nick)

        sql = """
            INSERT INTO testPosition
                (nickname, position, role_rate, win_rate)
            values
                (%s, %s, %s, %s)
            """
        cursor.execute(sql, (user_nick, position1, RoleRate1, WinRatio1))
        if position2 != '?':
            cursor.execute(sql, (user_nick, position2, RoleRate2, WinRatio2))


    a = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[3]/dl/dd[2]/a')
    driver.get(a.get_attribute('href'))
    driver.find_element_by_xpath('//*[@id="SummonerLayoutContent"]/div[3]/div/div/div[2]/div[1]/div/div[1]/div/div[2]').click()
    time.sleep(5)

    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')

    UserChampionList = soup.find('tbody', class_='Body')
    UserChampionListRows = UserChampionList.find_all('tr')

    i=1
    for UserChampionListRow in UserChampionListRows:
        ChampionName = UserChampionListRow.find(class_='ChampionName Cell').get_text().strip()
        try:  
            wins = int(UserChampionListRow.find(class_='Text Left').get_text()[:-1])
        except AttributeError:
            wins = 0

        try:  
            losses = int(UserChampionListRow.find(class_='Text Right').get_text()[:-1])
        except AttributeError:
            losses = 0

        Kill = float(UserChampionListRow.find(class_='Kill').get_text())
        Death = float(UserChampionListRow.find(class_='Death').get_text())
        Assist = float(UserChampionListRow.find(class_='Assist').get_text())

        try:
            sql = """
            INSERT INTO testCharacterInfo
                (most_index, nickname, character_name, wins, losses, kills, deaths, assist)
            values
                (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (i, user_nick, ChampionName, wins, losses, Kill, Death, Assist))
        except pymysql.IntegrityError:
            sql = """
            UPDATE testCharacterInfo SET character_name = %s, wins = %s, losses = %s, kills = %s, deaths = %s, assist = %s WHERE most_index = %s AND nickname = %s
            """
            cursor.execute(sql, (ChampionName, wins, losses, Kill, Death, Assist, i, user_nick))

        i=i+1

absadd = os.path.dirname( os.path.abspath( __file__ ) )
print(absadd)

load_wb = load_workbook(absadd+"/lol_list.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet']

host_name = 'gamehaeduo-db.c8xdbny5rkis.ap-northeast-2.rds.amazonaws.com' # DB 주소 세팅
username = 'gamehaeduo'
password = 'caugamehaeduo'
database_name = 'gamehaeduo'

db = pymysql.connect( # DB 연결
    host= host_name,
    port=3306,
    user= username,
    passwd= password,
    db= database_name,
    charset='utf8')
cursor = db.cursor()

driverpath = os.path.join(os.path.dirname(os.path.abspath(__file__)),"chromedriver.exe")
driver = webdriver.Chrome(executable_path=driverpath)
#만약 여기서 오류나면 자기 컴퓨터 절대경로로 바꿔주세용ㅎㅎ 
#예시)driver = webdriver.Chrome(r"C:\\Users\\이민희\\Downloads\\chromedriver.exe") 

count = 0
for col in load_ws.columns:
    count += 1
    if count == 1 or count ==3 :
        for row in range(1, load_ws.max_row):
            crawling(col[row].value)
    if count == 2:
        continue
    if count == 4:
        break

db.commit()

db.close()